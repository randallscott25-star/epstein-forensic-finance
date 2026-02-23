"""
EPSTEIN FORENSIC WORKBOOK v7
================================
DATABASE-DRIVEN REBUILD — Single source of truth.
All tabs now read from master_wire_ledger + bank_statement_transactions in the DB.
Published JSON (phase24) preserved as audit trail, not consumed.

Changes from v6.1:
  - DATA SOURCE: DB (master_wire_ledger table) replaces JSON
  - PROMOTED: Net-new wires from universal bank parser (Phase 5H)
  - NEW TAB 12: Multi-Bank Wire Ledger — 14 banks, wire/book transfers
  - NEW TAB 13: Bank Statement Coverage — heat map, 14 banks × 22 years
  - NEW TAB 14: Multi-Bank Summary — pipeline state, validation tiers
  - All 11 original tabs preserved with identical logic

Tabs (14):
  1. Executive Summary       — headline numbers, three-tier framework
  2. Extraction Phases       — 24-phase pipeline, running totals
  3. Money Flow Patterns     — how money moved through the network
  4. Shell Trust Hierarchy   — 4-tier trust structure with flows
  5. Master Wire Ledger      — all wires from master_wire_ledger table
  6. Above-Cap Verified      — wires above $10M ceiling
  7. Date Recovery           — same-amount different-date analysis
  8. Entity P&L              — who got what, inflows vs outflows
  9. Shell Network           — shell-involved wires with hop classification
  10. SAR Comparison          — vs FinCEN benchmarks
  11. Methodology             — bugs found, data sources, limitations
  12. Multi-Bank Wires        — ★ NEW: 14-bank wire/book ledger
  13. Bank Stmt Coverage      — ★ NEW: coverage heat map
  14. Multi-Bank Summary      — ★ NEW: pipeline + validation tiers

Requires: epstein_files.db with master_wire_ledger, bank_statement_transactions,
          verified_wires, fincen_transactions tables.

Usage:
  python forensic_workbook_v7.py 2>&1 | tee forensic_workbook_v7_output.txt
"""

import json, re, sys, sqlite3
from pathlib import Path
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════
HOME = Path.home()
REPORT_DIR = HOME / "epstein_reports"
REPORT_DIR.mkdir(exist_ok=True)
OUTFILE = "EPSTEIN_FORENSIC_WORKBOOK_v7.xlsx"
HDD_DIR = Path("/Volumes/My Book/epstein_project/reporting")
DB_PATHS = [
    Path("/Volumes/My Book/epstein_project/epstein_files.db"),
    Path.cwd() / "epstein_files.db",
    HOME / "epstein_files.db",
]
LEDGER_PATHS = [
    HOME / "epstein_project" / "reports" / "master_wire_ledger_phase24.json",
    REPORT_DIR / "master_wire_ledger_phase24.json",
    HOME / "master_wire_ledger_phase24.json",
    Path.cwd() / "master_wire_ledger_phase24.json",
]
SAR_TOTAL = 1_878_000_000
SAR_BENCHMARKS = {
    "JPMorgan Chase": {"amount": 1_100_000_000, "source": "Wyden Senate Finance Committee"},
    "Deutsche Bank": {"amount": 400_000_000, "source": "FinCEN SAR filings"},
    "BNY Mellon": {"amount": 378_000_000, "source": "FinCEN SAR filings"},
}

# ══════════════════════════════════════════════════════════════
# FINAL AUDITED NUMBERS (Phase 24)
# ══════════════════════════════════════════════════════════════
V2_20_BASE = 1_652_349_113
DATE_RECOVERED = 189_137_320
ABOVE_CAP = 120_575_938
GRAND_TOTAL = 1_964_229_742
TIER_CONSERVATIVE = 1_843_653_804
TIER_PUBLICATION = 1_964_229_742
TIER_EXPANDED = 1_956_153_971

# ══════════════════════════════════════════════════════════════
# ENTITY CLASSIFICATION — FIXED for underscore PROVEN names
# ══════════════════════════════════════════════════════════════
# Epstein-controlled entities (trusts, shells, personal accounts, attorneys)
EPSTEIN_ENTITIES_PATTERNS = [
    'southern trust', 'southern financial', 'haze trust', 'the haze trust',
    'butterfly trust', 'butterfly_trust', 'gratitude america', 'gratitude_america',
    'plan d', 'plan b', 'bv70', 'honeycomb', 'boothbay', 'caterpillar trust',
    'caterpillar_trust', '2017 caterpillar', 'valar', 'jege', 'lsj', 'lsje',
    'zorro', 'hyperion', 'nm property', 'razorback', 'lcp', 'red hook',
    'little st james', 'great st james', 'nes llc',
    'jeffrey epstein', 'epstein now', 'supernow', 'jeffery epstein',
    'mr epstein', 'epstein & co', 'jeffrey -', 'e stein',
    'ghislaine maxwell', 'jeepers', 'indyke', 'kahn', 'halperin',
    'edgarley', 'ellmax', 'nightfly', 'nautilus inc', 'maple inc',
    'laurel inc', 'cypress inc', 'terramar', 'air ghislaine',
    'thomas world air', 'f t real estate', 'financial strategy group',
    'financial_trust', 'max hotel', "michelle's transportation", 'liquid funding',
    'hbrk', 'rytanee', 'financial trust',
    # Underscore variants from PROVEN pipeline
    'southern_trust', 'southern_financial', 'haze_trust',
]

BANK_PATTERNS = [
    'deutsche', 'jpmorgan', 'chase bank', 'citibank', 'hsbc',
    'bank of america', 'ubs ', 'morgan stanley', 'credit suisse',
    'barclays', 'pershing', 'bny mellon', 'bny_mellon',
    'national financial', 'charles schwab', 'first bank pr',
]

# Shell entity display names for shell tabs
SHELL_ENTITIES = {
    "Southern Trust Company Inc.", "Southern Trust Company Inc. (Checking)",
    "Southern Financial LLC", "Southern Financial LLC (Checking)",
    "Southern Financial LLC (DBAGNY)",
    "The Haze Trust (DBAGNY)", "The Haze Trust (Checking)",
    "Butterfly Trust", "Gratitude America Ltd.", "Gratitude America MMDA",
    "Gratitude America Ltd. (Morgan Stanley/Citibank)",
    "Gratitude America Ltd. (First Bank PR)",
    "Plan D LLC", "BV70 LLC", "Honeycomb Partners LP",
    "Boothbay Multi-Strategy Fund", "The 2017 Caterpillar Trust",
    "Valar Global", "JEGE LLC", "LSJ Employees LLC",
    "Zorro Trust", "Hyperion Air Inc", "NM Property LLC",
    "Razorback LLC", "NES LLC", "Jeepers Inc.",
    "Jeepers Inc. (DB Brokerage)", "PLAN D LLC",
    "F T Real Estate Inc", "Financial Strategy Group Ltd",
    "Edgarley LLC", "Ellmax LLC", "Nightfly Inc",
    "Maple Inc", "Laurel Inc", "Cypress Inc", "Nautilus Inc",
    "Terramar Project Inc", "Thomas World Air LLC",
    "Max Hotel Services Corp", "Michelle's Transportation Company",
    "Liquid Funding Ltd", "HBRK Associates Inc", "Rytanee LLC",
    "Financial Trust Company", "LCP Company",
    # Underscore variants
    "SOUTHERN_TRUST", "SOUTHERN_FINANCIAL", "HAZE_TRUST",
    "BUTTERFLY_TRUST", "GRATITUDE_AMERICA", "FINANCIAL_TRUST",
}


def classify_entity(name):
    if not name or len(str(name).strip()) < 3:
        return 'UNKNOWN'
    n = str(name).lower().strip().replace('_', ' ')
    for b in BANK_PATTERNS:
        if b in n:
            return 'BANK/CUSTODIAN'
    for ep in EPSTEIN_ENTITIES_PATTERNS:
        if ep in n:
            return 'EPSTEIN ENTITY'
    return 'EXTERNAL PARTY'


def is_shell(name):
    if not name:
        return False
    n = str(name).strip()
    if n in SHELL_ENTITIES:
        return True
    nl = n.lower().replace('_', ' ')
    for ep in EPSTEIN_ENTITIES_PATTERNS:
        if ep in nl:
            return True
    return False


def flow_label(from_type, to_type):
    """Human-readable flow label for the money movement pattern."""
    labels = {
        ('EXTERNAL PARTY', 'EPSTEIN ENTITY'): 'MONEY IN — External funds entering Epstein network',
        ('EPSTEIN ENTITY', 'EPSTEIN ENTITY'): 'INTERNAL MOVE — Reshuffled between Epstein trusts/shells',
        ('EPSTEIN ENTITY', 'EXTERNAL PARTY'): 'MONEY OUT — Funds leaving to external beneficiaries',
        ('BANK/CUSTODIAN', 'EPSTEIN ENTITY'): 'BANK → SHELL — Custodian disbursement to Epstein entity',
        ('EPSTEIN ENTITY', 'BANK/CUSTODIAN'): 'SHELL → BANK — Epstein entity depositing to bank/custodian',
        ('EXTERNAL PARTY', 'EXTERNAL PARTY'): 'PASS-THROUGH — Between external parties (attorney, trust admin)',
        ('BANK/CUSTODIAN', 'EXTERNAL PARTY'): 'BANK → EXTERNAL — Custodian disbursement to non-Epstein party',
        ('EXTERNAL PARTY', 'BANK/CUSTODIAN'): 'EXTERNAL → BANK — External party depositing to bank',
        ('BANK/CUSTODIAN', 'BANK/CUSTODIAN'): 'INTERBANK — Between financial institutions',
    }
    return labels.get((from_type, to_type), f'{from_type} → {to_type}')


# ══════════════════════════════════════════════════════════════
# SHELL TRUST HIERARCHY
# ══════════════════════════════════════════════════════════════
HIERARCHY = [
    # (Entity, Tier, Role, Description)
    ("Southern Trust Company Inc.", 1, "PRIMARY HOLDING TRUST",
     "Main receiving entity for external investor funds. $151M+ inflow from Black, Rothschild, Narrow Holdings."),
    ("The 2017 Caterpillar Trust", 1, "INVESTMENT HOLDING TRUST",
     "Received $15M from Blockchain Capital venture funds."),
    ("The Haze Trust (DBAGNY)", 2, "DISTRIBUTION TRUST",
     "Redistributes funds: $32M → Southern Financial, $10M → Southern Trust, $5M → Southern Financial (DBAGNY)."),
    ("The Haze Trust (Checking)", 2, "ASSET RECEIPT TRUST",
     "Receives auction/art proceeds: $11.2M Sotheby's, $7.7M Christie's."),
    ("Southern Financial LLC", 2, "INVESTMENT PASS-THROUGH",
     "Receives from Tudor Futures ($13.5M), distributes to Coatue ($2M), Ito ($1M), Neoteny ($1M)."),
    ("Southern Financial LLC (Checking)", 2, "INTERNAL LEDGER",
     "Receives $32M from Haze Trust. Internal accounting position."),
    ("Jeepers Inc. (DB Brokerage)", 3, "PERSONAL FUNDING VEHICLE",
     "Funnels $51.9M directly to Epstein personal account (21 wires, brokerage-to-checking)."),
    ("Jeepers Inc.", 3, "SHELL FEEDER",
     "Feeds $6M to Jeepers Inc. (DB Brokerage) — one shell funding another."),
    ("Plan D LLC", 3, "DISBURSEMENT SHELL",
     "Paid $18M to Leon Black (4 wires: $8M, $5M, $3M, $2M). Outbound-only shell."),
    ("Gratitude America MMDA", 3, "CHARITABLE FACADE",
     "Distributed $6.3M: $5.5M to Morgan Stanley/Citibank, $300K to First Bank PR, small grants to medical charities."),
    ("NES LLC", 3, "MAXWELL FUNDING SHELL",
     "Paid $539K to Ghislaine Maxwell. Small operating disbursements (Pottery Barn, Visa)."),
    ("HALPERIN", 3, "ATTORNEY DISBURSEMENT",
     "Paid $3.2M to Leon Black through attorney channel."),
    ("KAHN", 3, "ATTORNEY PASS-THROUGH",
     "Richard Kahn: received $1.4M (Epstein, Pagano), paid out $9.3M (Paul Morris $8.5M, Tazia Smith $798K)."),
    ("Jeffrey Epstein NOW/SuperNow Account", 4, "PERSONAL ACCOUNT",
     "Terminal destination. Received $83.4M (Jeepers $51.9M, Kellerhals $23M, law firms $6M+). No outflows in ledger."),
    ("INDYKE", 4, "ESTATE ATTORNEY",
     "Darren Indyke: received $6.4M (Deutsche $5.8M), disbursed $313K (Goldberg, Visoski, Hanna)."),
    ("Hyperion Air Inc", 4, "AVIATION SHELL",
     "Aircraft-related entity. Small flows only ($196K in, $500K to BofA)."),
]

# ══════════════════════════════════════════════════════════════
# EXTRACTION PHASE LOG
# ══════════════════════════════════════════════════════════════
PHASES = [
    ("v2", "Core OCR Extraction", "Regex on 1.575M files across 19 datasets", 1204000000, 1204000000,
     "Amount-unique dedup", "Base extraction — $1.2B from raw OCR text"),
    ("BF adj", "Butterfly Trust Correction", "Manual audit of trust fund balances", -63000000, 1141000000,
     "Manual review", "Running balances mistakenly counted as wire transfers"),
    ("v3.2", "7-Layer Wire Expansion", "SWIFT MT103 + AUTH TO TRANSFER + bank corpus parsing", 237452186, 1378452186,
     "7-layer proximity filter", "Cross-bank dedup, context validation, IMAD matching"),
    ("14.5", "Known Entity Fund Flows", "fund_flows table matched against known entity list", 90936712, 1469388898,
     "Entity name match", "First structured table extraction from fund_flows_audited"),
    ("14.5B", "Balance Contamination Fix", "Removed running statement balances misidentified as wires", -53766217, 1415622681,
     "BUG FIX", "Statement balances ≠ wire amounts — $54M inflation removed"),
    ("15E", "Fund Flows Real Wires", "Strict wire indicators: Fedwire, IMAD, credit advice", 14000000, 1429622681,
     "Wire keyword required", "Only entries with explicit wire transfer language"),
    ("15F", "Redaction Recovery", "OCR fragments near [REDACTED] markers", 368170, 1429990851,
     "Proximity scan", "Partial dollar amounts visible adjacent to redactions"),
    ("16.1", "Transaction-Line Parser", "Structured line-by-line parsing of bank statements", 78547827, 1508538678,
     "Format validation", "Cleaned: removed $155M in brokerage/balance noise"),
    ("16.2", "Round-Wire Extractor", "Exact round-dollar amounts with wire context", 4975350, 1513514028,
     "Round-amount + context", "Conservative: only round amounts with clear wire language"),
    ("17", "Trust Transfers + DS8", "trust_transfers table + Dataset 8 financial records", 12720752, 1526234780,
     "Table structure", "New tables: trust_transfers, DS8 financial_hits"),
    ("18", "Full Category Sweep", "All remaining fund_flows_audited categories", 3854313, 1530089093,
     "Category filter", "Swept wire_transfer, fedwire, check, ach categories"),
    ("19", "Audited PROVEN Fix", "Fixed self-dedup bug in fund_flows_audited", 59524629, 1589613722,
     "BUG FIX", "Table was checking against itself — $59M real wires recovered"),
    ("20A", "Verified Wires (amount-new)", "verified_wires table — amounts not in prior phases", 53093926, 1642707648,
     "Court-exhibit verified", "185 wires with exhibit numbers, bates stamps, dates"),
    ("21A/B", "STRONG/MODERATE New Amounts", "fund_flows_audited STRONG+MODERATE tiers, wire-indicated", 9641465, 1652349113,
     "Wire indicator + tier", "Only truly new amounts not captured by any prior phase"),
    ("23", "Date-Aware Wire Census", "Master ledger: amount + entity + date signature", 191304691, 1843653804,
     "Date dedup", "95 same-amount different-date wires recovered — $189M"),
    ("24", "Above-Cap Verified Wires", "Court-exhibit verified wires above old $10M ceiling", 120575938, 1964229742,
     "Court-exhibit verified", "8 wires: Kellerhals $23M, Narrow Holdings $20M, Black $15M, etc."),
]

# ══════════════════════════════════════════════════════════════
# BUGS FOUND
# ══════════════════════════════════════════════════════════════
BUGS_FOUND = [
    ("14.5B", "Balance Contamination", "-$53.8M",
     "Running statement balances (e.g., 'ENDING BALANCE: $4,200,000') were extracted as wire amounts.",
     "Added is_balance flag; excluded from all subsequent phases."),
    ("16.1", "Brokerage Noise", "-$155M",
     "Securities language ('mutual fund', 'portfolio value', 'unrealized gain') mixed with wire data.",
     "Added BROKERAGE_NOISE regex filter; removed from 16.1 extraction."),
    ("19", "Self-Dedup Bug", "+$59.5M",
     "fund_flows_audited was checking new entries against ITSELF, causing legitimate wires to be skipped.",
     "Fixed to check only against prior-phase JSON amounts. $59M in real wires recovered."),
    ("20D", "Amount-Only Dedup", "+$115M",
     "Same dollar amount between DIFFERENT entity pairs was collapsed to one entry.",
     "Entity-pair aware dedup recovered $115M (before chain-hop filter)."),
    ("22", "Chain-Hop Inflation", "-$311M",
     "$10M from Black → Southern Trust → Southern Financial → Boothbay counted 4 times as $40M.",
     "Built INTERNAL entity taxonomy. Identified and excluded INTERNAL→INTERNAL chain hops."),
    ("22", "Cross-Table Name Duplication", "Risk flagged",
     "Same wire in verified_wires as 'Leon & Debra Black' and fund_flows as 'LEON_BLACK'.",
     "Three-tier confidence framework. Amount-only dedup for conservative tier."),
    ("22", "Statement Noise as Entities", "-$59M",
     "'Balance Transfers', 'Gift Cards', 'Dear Jeff', 'Schedule A' classified as wire entities.",
     "Built STATEMENT_NOISE filter. 188 entries ($59M) removed."),
    ("23", "Date-Blind Dedup", "+$189M",
     "Four $10M Black wires on different dates (2013-12-18, 2015-12-18, etc.) counted as ONE.",
     "Date-aware census: amount+entity+date. 95 extra instances recovered."),
    ("24", "Arbitrary Cap + Custodian Suffix", "+$120.6M / -$113.4M",
     "$10M cap excluded 8 verified wires. '(DBAGNY)' suffix misclassified entities as bank.",
     "Cap raised for verified tier. Custodian audit: 36/37 'bank' wires were chain hops."),
]

# ══════════════════════════════════════════════════════════════
# STYLING
# ══════════════════════════════════════════════════════════════
MONEY = '$#,##0'
MONEY_NEG = '$#,##0;($#,##0);"-"'
PCT = '0.0%'
HF = Font(bold=True, color='FFFFFF', size=11, name='Arial')
BF = Font(name='Arial', size=10)
SF = Font(name='Arial', size=9)
BOLD_DATA = Font(name='Arial', bold=True, size=10, color='1B2A4A')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1B2A4A')
SUBTITLE = Font(name='Arial', bold=True, size=11, color='2C3E6B')
GREEN_NUM = Font(name='Arial', bold=True, size=11, color='1B6B3A')
RED_NUM = Font(name='Arial', size=10, color='C0392B')
WARN_FONT = Font(name='Arial', size=9, color='CC0000', italic=True)

TB = Border(bottom=Side(style='thin', color='D0D0D0'))
THICK_BOT = Border(bottom=Side(style='medium', color='1B2A4A'))
Z1 = PatternFill('solid', fgColor='F2F7FB')
Z2 = PatternFill()

DARK_HDR = PatternFill('solid', fgColor='1F4E79')
GREEN_HDR = PatternFill('solid', fgColor='375623')
RED_HDR = PatternFill('solid', fgColor='C00000')
GOLD_HDR = PatternFill('solid', fgColor='7F6000')
TEAL_HDR = PatternFill('solid', fgColor='1A5276')
PURPLE_HDR = PatternFill('solid', fgColor='6C3483')

MONEY_IN_FILL = PatternFill('solid', fgColor='D5F5E3')    # green — money entering
INTERNAL_FILL = PatternFill('solid', fgColor='FCE4D6')     # orange — internal reshuffling
MONEY_OUT_FILL = PatternFill('solid', fgColor='FADBD8')    # red — money leaving
BANK_FILL = PatternFill('solid', fgColor='D6EAF8')         # blue — bank intermediary
PASS_FILL = PatternFill('solid', fgColor='F2F3F4')         # grey — pass-through
PROVEN_FILL = PatternFill('solid', fgColor='C6EFCE')
ABOVE_CAP_FILL = PatternFill('solid', fgColor='F9E79F')
BUG_POS = PatternFill('solid', fgColor='D5F5E3')
BUG_NEG = PatternFill('solid', fgColor='FADBD8')
TOTAL_FILL = PatternFill('solid', fgColor='C6EFCE')
SECTION = PatternFill('solid', fgColor='D9E2F3')
DISCLAIMER_FILL = PatternFill('solid', fgColor='FFFFCC')

CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

DISCLAIMER = (
    "SCOPE & LIMITATIONS: This analysis does not constitute an audit, examination, "
    "or review performed in accordance with GAAS, GAGAS, or AICPA SSFS No. 1. "
    "Dollar figures represent amounts identified through automated text extraction "
    "and cross-referenced against court exhibits. Verification tiers are analyst "
    "assessments, not audit opinions."
)

def sanitize(text):
    if not text:
        return ''
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', str(text)).replace('\n', ' ').strip()

def make_header(ws, cols, fill, row=1):
    for c, val in enumerate(cols, 1):
        cl = ws.cell(row, c, val)
        cl.font = HF; cl.fill = fill; cl.alignment = CTR; cl.border = TB
    ws.row_dimensions[row].height = 36
    ws.freeze_panes = f'A{row + 1}'

def add_disclaimer(ws, col_count):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(col_count, 12))
    cl = ws.cell(1, 1, DISCLAIMER)
    cl.font = WARN_FONT; cl.fill = DISCLAIMER_FILL; cl.alignment = LEFT_WRAP
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = 'A3'

def widths(ws, w_list):
    for i, w in enumerate(w_list):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

def zebra(ws, r, max_col, base_row=2):
    fill = Z1 if (r - base_row) % 2 == 0 else Z2
    for c in range(1, max_col + 1):
        cl = ws.cell(r, c)
        if not cl.fill or cl.fill.fgColor is None or cl.fill.fgColor.rgb in ('00000000', None):
            cl.fill = fill
        cl.border = TB
        if not cl.font or cl.font.name is None:
            cl.font = BF

def find_db():
    for p in DB_PATHS:
        if p.exists():
            return str(p)
    return None

def load_master_ledger():
    """Load from DB first (single source of truth), fallback to JSON."""
    db_path = find_db()
    if db_path:
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT * FROM master_wire_ledger ORDER BY amount DESC")
            rows = cur.fetchall()
            cols = [d[0] for d in cur.description]
            data = []
            for r in rows:
                d = dict(zip(cols, r))
                d['amount'] = float(d.get('amount', 0) or 0)
                d['entity_from'] = d.get('entity_from', '') or ''
                d['entity_to'] = d.get('entity_to', '') or ''
                d['source'] = d.get('source', '') or ''
                d['date'] = d.get('date', '') or ''
                d['exhibit'] = d.get('exhibit', '') or ''
                d['bates'] = d.get('bates', '') or ''
                d['is_date_recovery'] = bool(d.get('is_date_recovery', 0))
                d['was_in_amount_dedup'] = bool(d.get('was_in_amount_dedup', 0))
                data.append(d)
            conn.close()
            print(f"  Loaded from DB: {db_path} ({len(data)} entries, ${sum(d['amount'] for d in data):,.0f})")
            return data
        except Exception as e:
            print(f"  DB load failed ({e}), falling back to JSON...")
    for p in LEDGER_PATHS:
        if p.exists():
            with open(p) as f:
                data = json.load(f)
            print(f"  Loaded from JSON (fallback): {p} ({len(data)} entries)")
            return data
    print("  ❌ No data source found")
    return []


# ══════════════════════════════════════════════════════════════
# TAB 1: EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════
def tab_executive_summary(wb):
    print("  Tab 1: Executive Summary...")
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = "1B2A4A"

    ws.merge_cells('A1:H1')
    ws.cell(1, 1, "EPSTEIN EFTA FORENSIC FINANCIAL EXTRACTION").font = TITLE_FONT
    ws.merge_cells('A2:H2')
    ws.cell(2, 1, "Pro Bono Analysis — 19 Datasets · 1.575M Files · 24 Extraction Phases · February 2025"
            ).font = Font(name='Arial', size=10, color='7F8C8D', italic=True)

    ws.merge_cells('A4:B4')
    ws.cell(4, 1, "TOTAL FINANCIAL ACTIVITY EXTRACTED").font = SUBTITLE
    ws.cell(4, 3, GRAND_TOTAL).font = Font(name='Arial', bold=True, size=16, color='1B6B3A')
    ws.cell(4, 3).number_format = MONEY

    ws.merge_cells('A5:B5')
    ws.cell(5, 1, "FinCEN SAR Benchmark").font = BF
    ws.cell(5, 3, SAR_TOTAL).font = BOLD_DATA
    ws.cell(5, 3).number_format = MONEY

    ws.merge_cells('A6:B6')
    ws.cell(6, 1, "Extraction Coverage (% of SAR)").font = BF
    ws.cell(6, 3).value = '=C4/C5'
    ws.cell(6, 3).number_format = PCT
    ws.cell(6, 3).font = GREEN_NUM

    r = 8
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1, "THREE-TIER CONFIDENCE FRAMEWORK").font = SUBTITLE
    r += 1
    for c, val in enumerate(["Tier", "Name", "Extracted Amount", "% of SAR", "What's Included", "Duplication Risk"], 1):
        cl = ws.cell(r, c, val); cl.font = HF; cl.fill = DARK_HDR; cl.alignment = CTR
    r += 1
    tier_data = [
        ("1", "Conservative", TIER_CONSERVATIVE, TIER_CONSERVATIVE / SAR_TOTAL,
         "v2-20 amount-unique + Phase 23 date recovery", "Zero — most restrictive dedup"),
        ("2", "Publication ★", TIER_PUBLICATION, TIER_PUBLICATION / SAR_TOTAL,
         "Tier 1 + 8 above-cap court-verified wires ($120.6M)", "Zero — all verified with exhibit numbers"),
        ("3", "Expanded", TIER_EXPANDED, TIER_EXPANDED / SAR_TOTAL,
         "Tier 2 + PROVEN entity expansion", "Minor — possible cross-table name overlap"),
    ]
    for i, (tier, name, amt, pct, desc, risk) in enumerate(tier_data):
        ws.cell(r, 1, tier).alignment = CTR
        ws.cell(r, 2, name).font = BOLD_DATA if '★' in name else BF
        ws.cell(r, 3, amt).number_format = MONEY
        ws.cell(r, 4, pct).number_format = PCT
        ws.cell(r, 5, desc).font = SF
        ws.cell(r, 6, risk).font = SF
        fill = TOTAL_FILL if '★' in name else (Z1 if i % 2 == 0 else Z2)
        for c in range(1, 7):
            ws.cell(r, c).fill = fill; ws.cell(r, c).border = TB
        r += 1

    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1, "WHY THE TOTAL EXCEEDS THE SAR BENCHMARK (>100%)").font = SUBTITLE
    r += 1
    for line in [
        "The SAR benchmark ($1.878B) = only transactions banks flagged as SUSPICIOUS.",
        "The EFTA corpus = the COMPLETE financial record, including legitimate activity:",
        "   · Sotheby's auction proceeds ($11.2M) — art sales, not suspicious",
        "   · Tudor Futures investment return ($12.8M) — normal hedge fund activity",
        "   · Kellerhals law firm settlement ($23M) — legal proceedings",
        "   · Blockchain Capital VC investment ($10.5M) — venture capital",
        "",
        "Total financial flows SHOULD exceed the suspicious subset.",
        "Standard forensic accounting: SAR ⊂ Total Financial Activity.",
    ]:
        ws.merge_cells(f'A{r}:F{r}'); ws.cell(r, 1, line).font = SF; r += 1

    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1, "KEY FINDINGS").font = SUBTITLE
    r += 1
    for label, val in [
        ("Datasets processed", "19"), ("Files in EFTA corpus", "1,575,000"),
        ("Entities identified", "11,400,000"), ("Extraction phases", "24"),
        ("Bugs caught & fixed", "9"), ("Wires in master ledger", "382"),
        ("Unique dated wires", "258"), ("GitHub", "randallscott25-star/epstein-efoia-index"),
        ("Submitted to Sen. Barton", "February 18, 2025"),
    ]:
        ws.cell(r, 1, label).font = BF; ws.cell(r, 2, val).font = BOLD_DATA
        for c in (1, 2): ws.cell(r, c).border = TB
        r += 1

    widths(ws, [34, 26, 22, 14, 46, 36])


# ══════════════════════════════════════════════════════════════
# TAB 2: EXTRACTION PHASES
# ══════════════════════════════════════════════════════════════
def tab_extraction_phases(wb):
    print("  Tab 2: Extraction Phases...")
    ws = wb.create_sheet("Extraction Phases")
    ws.sheet_properties.tabColor = "2E86C1"
    cols = ["Phase", "What We Did", "How We Did It", "Amount Added/Removed",
            "Running Total", "% of SAR Recovered", "Quality Check", "Key Finding"]
    make_header(ws, cols, TEAL_HDR)
    for i, (phase, desc, method, amt, running, gate, finding) in enumerate(PHASES):
        r = i + 2
        ws.cell(r, 1, phase).font = BOLD_DATA; ws.cell(r, 1).alignment = CTR
        ws.cell(r, 2, desc).font = BF
        ws.cell(r, 3, method).font = SF
        ws.cell(r, 4, amt).number_format = MONEY
        ws.cell(r, 5, running).number_format = MONEY
        ws.cell(r, 6, running / SAR_TOTAL).number_format = PCT
        ws.cell(r, 7, gate).font = SF
        ws.cell(r, 8, finding).font = SF
        if amt < 0:
            ws.cell(r, 4).font = RED_NUM
            for c in range(1, 9): ws.cell(r, c).fill = BUG_NEG; ws.cell(r, c).border = TB
        elif 'BUG FIX' in gate:
            ws.cell(r, 4).font = GREEN_NUM
            for c in range(1, 9): ws.cell(r, c).fill = BUG_POS; ws.cell(r, c).border = TB
        else:
            zebra(ws, r, 8)
    r = len(PHASES) + 2
    ws.cell(r, 1, "FINAL").font = Font(bold=True, name='Arial', size=11)
    ws.cell(r, 4, GRAND_TOTAL).number_format = MONEY
    ws.cell(r, 4).font = Font(bold=True, name='Arial', size=11, color='1B6B3A')
    ws.cell(r, 5, GRAND_TOTAL).number_format = MONEY
    ws.cell(r, 5).font = Font(bold=True, name='Arial', size=11, color='1B6B3A')
    ws.cell(r, 6, GRAND_TOTAL / SAR_TOTAL).number_format = PCT; ws.cell(r, 6).font = GREEN_NUM
    for c in range(1, 9):
        ws.cell(r, c).fill = TOTAL_FILL
        ws.cell(r, c).border = Border(top=Side(style='medium', color='1B2A4A'), bottom=Side(style='medium', color='1B2A4A'))
    widths(ws, [10, 28, 46, 18, 18, 14, 22, 54])
    add_disclaimer(ws, 8)


# ══════════════════════════════════════════════════════════════
# TAB 3: MONEY FLOW PATTERNS ★ NEW
# ══════════════════════════════════════════════════════════════
def tab_money_flow_patterns(wb, ledger):
    print("  Tab 3: Money Flow Patterns...")
    ws = wb.create_sheet("Money Flow Patterns")
    ws.sheet_properties.tabColor = "E74C3C"

    ws.merge_cells('A1:H1')
    ws.cell(1, 1, "HOW THE MONEY MOVED: External Sources → Epstein Trusts → Banks → Shells → Beneficiaries"
            ).font = TITLE_FONT
    ws.merge_cells('A2:H2')
    ws.cell(2, 1,
        "This tab shows the laundering pattern: money enters from wealthy external donors (Black, Rothschild), "
        "gets redistributed through a network of trusts and shell companies (Southern Trust → Haze Trust → "
        "Southern Financial → Jeepers), and exits to beneficiaries or personal accounts."
    ).font = SF
    ws.row_dimensions[2].height = 48

    # SECTION A: Flow type summary
    r = 4
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(r, 1, "SECTION A: MONEY FLOW SUMMARY BY DIRECTION").font = SUBTITLE
    ws.cell(r, 1).fill = SECTION
    r += 1
    cols_a = ["Flow Direction", "What This Means", "# of Wires", "Total Amount",
              "% of Ledger", "Largest Single Wire", "Key Example"]
    for c, val in enumerate(cols_a, 1):
        cl = ws.cell(r, c, val); cl.font = HF; cl.fill = RED_HDR; cl.alignment = CTR
    r += 1

    # Compute flows
    flow_types = defaultdict(lambda: {'count': 0, 'total': 0, 'max_amt': 0, 'max_ex': '', 'examples': []})
    for w in ledger:
        ef, et = w.get('entity_from', ''), w.get('entity_to', '')
        eft, ett = classify_entity(ef), classify_entity(et)
        label = flow_label(eft, ett)
        flow_types[label]['count'] += 1
        flow_types[label]['total'] += w['amount']
        if w['amount'] > flow_types[label]['max_amt']:
            flow_types[label]['max_amt'] = w['amount']
            flow_types[label]['max_ex'] = f"{sanitize(ef)[:25]} → {sanitize(et)[:25]}"
        if len(flow_types[label]['examples']) < 2:
            flow_types[label]['examples'].append(f"{sanitize(ef)[:20]} → {sanitize(et)[:20]} (${w['amount']:,.0f})")

    total_ledger = sum(w['amount'] for w in ledger)
    flow_fills = {
        'MONEY IN': MONEY_IN_FILL, 'INTERNAL MOVE': INTERNAL_FILL,
        'MONEY OUT': MONEY_OUT_FILL, 'BANK': BANK_FILL, 'SHELL': BANK_FILL,
        'INTERBANK': PASS_FILL, 'PASS-THROUGH': PASS_FILL, 'EXTERNAL': PASS_FILL,
    }

    for label, data in sorted(flow_types.items(), key=lambda x: -x[1]['total']):
        short = label.split('—')[0].strip() if '—' in label else label
        desc = label.split('—')[1].strip() if '—' in label else ''
        ws.cell(r, 1, short).font = BOLD_DATA
        ws.cell(r, 2, desc).font = SF
        ws.cell(r, 3, data['count']).alignment = CTR
        ws.cell(r, 4, data['total']).number_format = MONEY
        ws.cell(r, 5, data['total'] / total_ledger if total_ledger else 0).number_format = PCT
        ws.cell(r, 6, data['max_amt']).number_format = MONEY
        ws.cell(r, 7, data['max_ex']).font = SF

        fill = PASS_FILL
        for key, f in flow_fills.items():
            if key in short:
                fill = f; break
        for c in range(1, 8):
            ws.cell(r, c).fill = fill; ws.cell(r, c).border = TB
        r += 1

    # Total row
    ws.cell(r, 1, "TOTAL LEDGER VALUE").font = Font(bold=True, name='Arial', size=11)
    ws.cell(r, 3, len(ledger)).alignment = CTR
    ws.cell(r, 4, total_ledger).number_format = MONEY
    ws.cell(r, 4).font = Font(bold=True, name='Arial', size=11, color='1B6B3A')
    for c in range(1, 8):
        ws.cell(r, c).fill = TOTAL_FILL; ws.cell(r, c).border = THICK_BOT
    r += 2

    # SECTION B: Detailed wire list with flow classification
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(r, 1, "SECTION B: EVERY WIRE CLASSIFIED BY FLOW DIRECTION").font = SUBTITLE
    ws.cell(r, 1).fill = SECTION
    r += 1
    cols_b = ["Flow Direction", "Money Source (From)", "Money Destination (To)",
              "Wire Amount", "Date", "Source Classification", "From Entity Type", "To Entity Type"]
    for c, val in enumerate(cols_b, 1):
        cl = ws.cell(r, c, val); cl.font = HF; cl.fill = DARK_HDR; cl.alignment = CTR
    r += 1

    for w in sorted(ledger, key=lambda x: -x['amount']):
        ef, et = w.get('entity_from', ''), w.get('entity_to', '')
        eft, ett = classify_entity(ef), classify_entity(et)
        label = flow_label(eft, ett)
        short = label.split('—')[0].strip() if '—' in label else label

        ws.cell(r, 1, short).font = BOLD_DATA
        ws.cell(r, 2, sanitize(ef))
        ws.cell(r, 3, sanitize(et))
        ws.cell(r, 4, w['amount']).number_format = MONEY
        ws.cell(r, 5, w.get('date', '') or '')
        ws.cell(r, 6, w.get('source', ''))
        ws.cell(r, 7, eft).alignment = CTR
        ws.cell(r, 8, ett).alignment = CTR

        fill = PASS_FILL
        for key, f in flow_fills.items():
            if key in short: fill = f; break
        ws.cell(r, 1).fill = fill
        for c in range(2, 9):
            ws.cell(r, c).fill = Z1 if r % 2 == 0 else Z2
            ws.cell(r, c).border = TB; ws.cell(r, c).font = BF
        ws.cell(r, 1).border = TB
        r += 1

    widths(ws, [22, 38, 38, 16, 12, 18, 16, 16])


# ══════════════════════════════════════════════════════════════
# TAB 4: SHELL TRUST HIERARCHY ★ NEW
# ══════════════════════════════════════════════════════════════
def tab_shell_hierarchy(wb, ledger):
    print("  Tab 4: Shell Trust Hierarchy...")
    ws = wb.create_sheet("Shell Trust Hierarchy")
    ws.sheet_properties.tabColor = "8E44AD"

    ws.merge_cells('A1:I1')
    ws.cell(1, 1, "EPSTEIN SHELL COMPANY & TRUST HIERARCHY — How Money Was Layered"
            ).font = TITLE_FONT
    ws.merge_cells('A2:I2')
    ws.cell(2, 1,
        "Tier 1 = holding trusts that receive external money. Tier 2 = pass-through trusts that redistribute. "
        "Tier 3 = operating shells that pay beneficiaries. Tier 4 = personal accounts (terminal destinations). "
        "Money flows DOWN the tiers: External donors → Tier 1 → Tier 2 → Tier 3 → Tier 4 personal accounts."
    ).font = SF
    ws.row_dimensions[2].height = 48

    # SECTION A: Hierarchy map
    r = 4
    ws.merge_cells(f'A{r}:I{r}')
    ws.cell(r, 1, "SECTION A: TRUST NETWORK STRUCTURE").font = SUBTITLE
    ws.cell(r, 1).fill = SECTION
    r += 1

    cols_h = ["Tier", "Entity Name", "Role in Network", "Total Money Received",
              "Total Money Sent", "Net Position", "# Sources", "# Destinations", "Description"]
    for c, val in enumerate(cols_h, 1):
        cl = ws.cell(r, c, val); cl.font = HF; cl.fill = PURPLE_HDR; cl.alignment = CTR
    r += 1

    # Compute actual in/out — case-insensitive matching
    entity_in = defaultdict(float)
    entity_out = defaultdict(float)
    entity_in_ct = defaultdict(int)
    entity_out_ct = defaultdict(int)
    for w in ledger:
        ef, et = w.get('entity_from', ''), w.get('entity_to', '')
        if ef: entity_out[ef.upper()] += w['amount']; entity_out_ct[ef.upper()] += 1
        if et: entity_in[et.upper()] += w['amount']; entity_in_ct[et.upper()] += 1

    tier_fills = {
        1: PatternFill('solid', fgColor='D5F5E3'),  # green — receiving
        2: PatternFill('solid', fgColor='FCE4D6'),  # orange — redistributing
        3: PatternFill('solid', fgColor='FADBD8'),  # red — disbursing
        4: PatternFill('solid', fgColor='D6EAF8'),  # blue — terminal
    }

    for entity, tier, role, desc in HIERARCHY:
        eu = entity.upper()
        total_in = entity_in.get(eu, 0)
        total_out = entity_out.get(eu, 0)
        in_ct = entity_in_ct.get(eu, 0)
        out_ct = entity_out_ct.get(eu, 0)
        net = total_in - total_out

        ws.cell(r, 1, f"TIER {tier}").alignment = CTR
        ws.cell(r, 1).font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        ws.cell(r, 1).fill = tier_fills[tier]
        ws.cell(r, 2, entity).font = BOLD_DATA
        ws.cell(r, 3, role).font = SF
        ws.cell(r, 4, total_in).number_format = MONEY
        ws.cell(r, 5, total_out).number_format = MONEY
        ws.cell(r, 6, net).number_format = MONEY_NEG
        ws.cell(r, 6).font = GREEN_NUM if net >= 0 else RED_NUM
        ws.cell(r, 7, in_ct).alignment = CTR
        ws.cell(r, 8, out_ct).alignment = CTR
        ws.cell(r, 9, desc).font = SF
        ws.row_dimensions[r].height = 36

        for c in range(2, 10):
            ws.cell(r, c).border = TB
            if c == 9:
                ws.cell(r, c).alignment = LEFT_WRAP
            elif c in (7, 8):
                ws.cell(r, c).alignment = CTR
        ws.cell(r, 1).border = TB
        r += 1

    # SECTION B: The money circuit
    r += 1
    ws.merge_cells(f'A{r}:I{r}')
    ws.cell(r, 1, "SECTION B: THE MONEY CIRCUIT — Tracing $10M Through the Network").font = SUBTITLE
    ws.cell(r, 1).fill = SECTION
    r += 1

    circuit = [
        ("STEP 1", "Black Family Partners", "Southern Trust Company Inc.", "$10,000,000",
         "EXTERNAL → TIER 1", "External investor deposits into primary holding trust",
         MONEY_IN_FILL),
        ("STEP 2", "Southern Trust Company Inc.", "The Haze Trust (DBAGNY)", "$10,000,000",
         "TIER 1 → TIER 2", "Primary trust redistributes to distribution trust at Deutsche Bank",
         INTERNAL_FILL),
        ("STEP 3", "The Haze Trust (DBAGNY)", "Southern Financial LLC", "$10,000,000",
         "TIER 2 → TIER 2", "Distribution trust moves to investment pass-through",
         INTERNAL_FILL),
        ("STEP 4", "Southern Financial LLC", "Boothbay Multi-Strategy", "$10,000,000",
         "TIER 2 → TIER 3", "Pass-through invests in fund (Boothbay = Epstein-linked hedge fund)",
         INTERNAL_FILL),
        ("STEP 5", "Boothbay Multi-Strategy", "Honeycomb Partners LP", "$10,000,000",
         "TIER 3 → TIER 3", "Fund-to-fund transfer within Epstein investment network",
         INTERNAL_FILL),
    ]
    cols_c = ["Step", "Money Source (From)", "Money Destination (To)", "Amount",
              "Flow Direction", "What This Step Accomplished", ""]
    for c, val in enumerate(cols_c, 1):
        cl = ws.cell(r, c, val); cl.font = HF; cl.fill = RED_HDR; cl.alignment = CTR
    r += 1

    for step, frm, to, amt, direction, explanation, fill in circuit:
        ws.cell(r, 1, step).font = BOLD_DATA; ws.cell(r, 1).alignment = CTR
        ws.cell(r, 2, frm).font = BF
        ws.cell(r, 3, to).font = BF
        ws.cell(r, 4, amt).font = BOLD_DATA
        ws.cell(r, 5, direction).font = SF; ws.cell(r, 5).alignment = CTR
        ws.cell(r, 6, explanation).font = SF
        for c in range(1, 7):
            ws.cell(r, c).fill = fill; ws.cell(r, c).border = TB
        r += 1

    ws.cell(r, 1, "").font = BF
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1,
        "RESULT: ONE $10M wire was counted 5 times ($50M) before Phase 22 chain-hop filter caught it. "
        "This pattern repeated for $8.5M, $8M, $7M, $5M, $3M amounts — inflating totals by $311M."
    ).font = Font(name='Arial', bold=True, size=10, color='C0392B')
    ws.cell(r, 1).alignment = LEFT_WRAP
    ws.row_dimensions[r].height = 36
    for c in range(1, 7):
        ws.cell(r, c).fill = PatternFill('solid', fgColor='FADBD8'); ws.cell(r, c).border = TB

    widths(ws, [10, 36, 36, 16, 18, 52, 2, 10, 52])


# ══════════════════════════════════════════════════════════════
# TAB 5: MASTER WIRE LEDGER
# ══════════════════════════════════════════════════════════════
def tab_master_ledger(wb, ledger):
    print("  Tab 5: Master Wire Ledger...")
    ws = wb.create_sheet("Master Wire Ledger")
    ws.sheet_properties.tabColor = "27AE60"

    cols = ["#", "Evidence Source", "Transaction Date", "Money Source (From)",
            "Money Destination (To)", "Wire Amount (USD)", "Court Exhibit/Reference",
            "Source Entity Type", "Destination Entity Type", "Flow Direction",
            "Recovered via Date Dedup?", "Above Prior $10M Cap?"]
    make_header(ws, cols, GREEN_HDR)

    for i, w in enumerate(sorted(ledger, key=lambda x: -x['amount'])):
        r = i + 2
        ef, et = sanitize(w.get('entity_from', '')), sanitize(w.get('entity_to', ''))
        eft, ett = classify_entity(ef), classify_entity(et)
        label = flow_label(eft, ett)
        short = label.split('—')[0].strip() if '—' in label else label

        ws.cell(r, 1, i + 1).alignment = CTR
        ws.cell(r, 2, w.get('source', ''))
        ws.cell(r, 3, w.get('date', '') or '')
        ws.cell(r, 4, ef)
        ws.cell(r, 5, et)
        ws.cell(r, 6, w['amount']).number_format = MONEY
        ws.cell(r, 7, sanitize(w.get('exhibit', '') or ''))
        ws.cell(r, 8, eft).alignment = CTR
        ws.cell(r, 9, ett).alignment = CTR
        ws.cell(r, 10, short)
        ws.cell(r, 11, 'YES' if w.get('is_date_recovery') and w.get('was_in_amount_dedup') else '').alignment = CTR
        ws.cell(r, 12, 'YES' if w['amount'] > 10_000_000 else '').alignment = CTR

        src = w.get('source', '')
        tier_fill = PROVEN_FILL if src == 'verified_wires' else (PatternFill('solid', fgColor='BDD7EE') if 'PROVEN' in src else None)
        if w['amount'] > 10_000_000:
            tier_fill = ABOVE_CAP_FILL
        fill = Z1 if i % 2 == 0 else Z2
        for c in range(1, 13):
            cl = ws.cell(r, c)
            cl.fill = tier_fill if tier_fill and c == 2 else fill
            cl.border = TB
            if not cl.font or cl.font.name is None: cl.font = BF

    r = len(ledger) + 2
    ws.cell(r, 5, "TOTAL").font = Font(bold=True, name='Arial', size=11)
    ws.cell(r, 6, sum(w['amount'] for w in ledger)).number_format = MONEY
    ws.cell(r, 6).font = Font(bold=True, name='Arial', size=11, color='1B6B3A')
    for c in range(1, 13):
        ws.cell(r, c).fill = TOTAL_FILL
        ws.cell(r, c).border = Border(top=Side(style='medium', color='375623'))

    widths(ws, [6, 18, 12, 38, 38, 16, 18, 16, 16, 24, 14, 14])
    add_disclaimer(ws, 12)
    print(f"    {len(ledger)} wires, ${sum(w['amount'] for w in ledger):,.0f}")


# ══════════════════════════════════════════════════════════════
# TAB 6: ABOVE-CAP VERIFIED
# ══════════════════════════════════════════════════════════════
def tab_above_cap(wb, ledger):
    print("  Tab 6: Above-Cap Verified Wires...")
    ws = wb.create_sheet("Above-Cap Verified")
    ws.sheet_properties.tabColor = "F39C12"
    ws.merge_cells('A1:G1')
    ws.cell(1, 1, "PHASE 24: 8 Court-Verified Wires Above the $10M Safety Cap").font = SUBTITLE
    ws.merge_cells('A2:G2')
    ws.cell(2, 1,
        "Phases v2-23 capped extraction at $10M to prevent balance contamination. Phase 24 lifted "
        "the cap for verified_wires entries (court-exhibit authenticated with dates and bates numbers). "
        "All 8 are REAL external transactions — not internal reshuffling."
    ).font = SF
    ws.row_dimensions[2].height = 48

    cols = ["Transaction Date", "Money Source (From)", "Money Destination (To)",
            "Wire Amount (USD)", "Court Exhibit", "What This Transaction Was", "Flow Direction"]
    make_header(ws, cols, GOLD_HDR, row=4)
    ws.freeze_panes = 'A5'

    above = sorted([w for w in ledger if w['amount'] > 10_000_000], key=lambda x: -x['amount'])
    sig = {
        23000000: "Law firm settlement payment to Epstein personal account",
        20000000: "External investor (Elysium/Narrow Holdings) depositing into holding trust",
        15000000: "Leon Black personal funds deposited into holding trust",
        14999980: "Rothschild family investment deposited into holding trust",
        13000000: "Black Family Partners LP depositing into holding trust via Apollo",
        12826541: "Tudor Futures hedge fund — investment return to Southern Financial",
        11249417: "Sotheby's auction house — art sale proceeds to Haze Trust",
        10500000: "Blockchain Capital — venture capital investment to Caterpillar Trust",
    }
    for i, w in enumerate(above):
        r = i + 5
        ef, et = sanitize(w.get('entity_from', '')), sanitize(w.get('entity_to', ''))
        eft, ett = classify_entity(ef), classify_entity(et)
        ws.cell(r, 1, w.get('date', ''))
        ws.cell(r, 2, ef)
        ws.cell(r, 3, et)
        ws.cell(r, 4, w['amount']).number_format = MONEY
        ws.cell(r, 4).font = Font(name='Arial', bold=True, size=11, color='1B6B3A')
        ws.cell(r, 5, sanitize(w.get('exhibit', '') or ''))
        ws.cell(r, 6, sig.get(int(w['amount']), ''))
        ws.cell(r, 7, f"{eft} → {ett}")
        for c in range(1, 8):
            ws.cell(r, c).fill = ABOVE_CAP_FILL; ws.cell(r, c).border = TB
            if c != 4: ws.cell(r, c).font = BF
    r = len(above) + 5
    ws.cell(r, 3, "TOTAL").font = Font(bold=True, name='Arial', size=11)
    ws.cell(r, 4, sum(w['amount'] for w in above)).number_format = MONEY
    ws.cell(r, 4).font = Font(bold=True, name='Arial', size=12, color='1B6B3A')
    for c in range(1, 8):
        ws.cell(r, c).fill = TOTAL_FILL; ws.cell(r, c).border = Border(top=Side(style='medium', color='7F6000'))
    widths(ws, [14, 36, 36, 18, 16, 52, 24])


# ══════════════════════════════════════════════════════════════
# TAB 7: DATE RECOVERY
# ══════════════════════════════════════════════════════════════
def tab_date_recovery(wb, ledger):
    print("  Tab 7: Date Recovery Analysis...")
    ws = wb.create_sheet("Date Recovery")
    ws.sheet_properties.tabColor = "8E44AD"
    ws.merge_cells('A1:G1')
    ws.cell(1, 1, "PHASE 23: Same Dollar Amount, Different Dates = Different Wires").font = SUBTITLE
    ws.merge_cells('A2:G2')
    ws.cell(2, 1,
        "Phases v2-20 deduplicated by amount only. This killed legitimate repeat wires: four separate "
        "$10M wires from Black Family on different dates (2013-12-18, 2015-12-17, 2015-12-18, 2015-12-30) "
        "were counted as ONE. Phase 23 used amount+entity+date signatures to recover 95 extra instances."
    ).font = SF
    ws.row_dimensions[2].height = 48

    amt_groups = defaultdict(list)
    for w in ledger:
        if w.get('was_in_amount_dedup'):
            amt_groups[f"{w['amount']:.2f}"].append(w)
    multi = {k: v for k, v in amt_groups.items() if len(v) > 1}
    multi_sorted = sorted(multi.items(), key=lambda x: -float(x[0]) * len(x[1]))

    cols = ["Wire Amount", "# Separate Instances", "Extra Value Recovered",
            "Transaction Date", "Money Source (From)", "Money Destination (To)", "Evidence Source"]
    make_header(ws, cols, PURPLE_HDR, row=4)
    ws.freeze_panes = 'A5'

    r = 5; total_extra = 0
    for amt_key, wires in multi_sorted[:30]:
        amt = float(amt_key)
        extras = len(wires) - 1
        extra_val = amt * extras
        total_extra += extra_val
        ws.cell(r, 1, amt).number_format = MONEY; ws.cell(r, 1).font = BOLD_DATA
        ws.cell(r, 2, len(wires)).alignment = CTR; ws.cell(r, 2).font = BOLD_DATA
        ws.cell(r, 3, extra_val).number_format = MONEY
        ws.cell(r, 3).font = Font(name='Arial', bold=True, size=10, color='1B6B3A')
        for c in range(1, 8):
            ws.cell(r, c).fill = PatternFill('solid', fgColor='E8DAEF'); ws.cell(r, c).border = TB
        r += 1
        for w in sorted(wires, key=lambda x: x.get('date', '') or 'zzz'):
            ws.cell(r, 4, w.get('date', 'no date') or 'no date').font = SF
            ws.cell(r, 5, sanitize(w.get('entity_from', ''))[:40]).font = SF
            ws.cell(r, 6, sanitize(w.get('entity_to', ''))[:40]).font = SF
            ws.cell(r, 7, w.get('source', '')).font = SF
            zebra(ws, r, 7, 5)
            r += 1

    r += 1
    ws.cell(r, 1, "TOTAL RECOVERED").font = Font(bold=True, name='Arial', size=11)
    ws.cell(r, 3, total_extra).number_format = MONEY
    ws.cell(r, 3).font = Font(bold=True, name='Arial', size=12, color='1B6B3A')
    for c in range(1, 8):
        ws.cell(r, c).fill = TOTAL_FILL; ws.cell(r, c).border = Border(top=Side(style='medium', color='6C3483'))
    widths(ws, [16, 14, 18, 14, 38, 38, 18])


# ══════════════════════════════════════════════════════════════
# TAB 8: ENTITY P&L
# ══════════════════════════════════════════════════════════════
def tab_entity_pl(wb, ledger):
    print("  Tab 8: Entity P&L...")
    ws = wb.create_sheet("Entity P&L")
    ws.sheet_properties.tabColor = "1ABC9C"
    entity_flows = defaultdict(lambda: {'inflow': 0, 'outflow': 0, 'in_ct': 0, 'out_ct': 0,
                                         'sources': set(), 'type': 'UNKNOWN'})
    for w in ledger:
        ef, et = sanitize(w.get('entity_from', '')), sanitize(w.get('entity_to', ''))
        if ef and len(ef) > 2:
            entity_flows[ef]['outflow'] += w['amount']; entity_flows[ef]['out_ct'] += 1
            entity_flows[ef]['sources'].add(w.get('source', '')); entity_flows[ef]['type'] = classify_entity(ef)
        if et and len(et) > 2:
            entity_flows[et]['inflow'] += w['amount']; entity_flows[et]['in_ct'] += 1
            entity_flows[et]['sources'].add(w.get('source', '')); entity_flows[et]['type'] = classify_entity(et)

    cols = ["Entity Name", "Entity Type", "Total Money Received (USD)", "Total Money Sent (USD)",
            "Net Position (USD)", "# Wires In", "# Wires Out", "Epstein Shell?", "Evidence Sources"]
    make_header(ws, cols, GREEN_HDR)
    sorted_ents = sorted(entity_flows.items(), key=lambda x: x[1]['inflow'] + x[1]['outflow'], reverse=True)
    for i, (entity, data) in enumerate(sorted_ents):
        r = i + 2
        net = data['inflow'] - data['outflow']
        shell = is_shell(entity)
        ws.cell(r, 1, entity)
        ws.cell(r, 2, data['type']).alignment = CTR
        ws.cell(r, 3, data['inflow']).number_format = MONEY
        ws.cell(r, 4, data['outflow']).number_format = MONEY
        ws.cell(r, 5, net).number_format = MONEY_NEG
        ws.cell(r, 5).font = GREEN_NUM if net >= 0 else RED_NUM
        ws.cell(r, 6, data['in_ct']).alignment = CTR
        ws.cell(r, 7, data['out_ct']).alignment = CTR
        ws.cell(r, 8, 'YES' if shell else '').alignment = CTR
        ws.cell(r, 9, ', '.join(sorted(s for s in data['sources'] if s))).font = SF
        fill = Z1 if i % 2 == 0 else Z2
        for c in range(1, 10):
            ws.cell(r, c).fill = fill; ws.cell(r, c).border = TB
            if not ws.cell(r, c).font or ws.cell(r, c).font.name is None: ws.cell(r, c).font = BF
        if shell: ws.cell(r, 8).fill = PatternFill('solid', fgColor='FFF3E0')
        type_fills = {'EPSTEIN ENTITY': PatternFill('solid', fgColor='FCE4EC'),
                      'EXTERNAL PARTY': PatternFill('solid', fgColor='D5F5E3'),
                      'BANK/CUSTODIAN': PatternFill('solid', fgColor='D6EAF8')}
        if data['type'] in type_fills: ws.cell(r, 2).fill = type_fills[data['type']]
    widths(ws, [40, 16, 18, 18, 18, 10, 10, 12, 34])
    add_disclaimer(ws, 9)
    print(f"    {len(sorted_ents)} entities")


# ══════════════════════════════════════════════════════════════
# TAB 9: SHELL NETWORK
# ══════════════════════════════════════════════════════════════
def tab_shell_network(wb, ledger):
    print("  Tab 9: Shell Network...")
    ws = wb.create_sheet("Shell Network")
    ws.sheet_properties.tabColor = "E74C3C"
    cols = ["Money Source (From)", "Money Destination (To)", "Wire Amount (USD)",
            "Transaction Date", "Source Type", "Destination Type",
            "Shell-to-Shell?", "Flow Direction", "Evidence Source"]
    make_header(ws, cols, RED_HDR)
    rows = []
    for w in ledger:
        ef, et = sanitize(w.get('entity_from', '')), sanitize(w.get('entity_to', ''))
        f_shell, t_shell = is_shell(ef), is_shell(et)
        if f_shell or t_shell:
            eft, ett = classify_entity(ef), classify_entity(et)
            label = flow_label(eft, ett)
            short = label.split('—')[0].strip() if '—' in label else label
            s2s = 'YES' if f_shell and t_shell else ''
            rows.append((ef, et, w['amount'], w.get('date', ''), eft, ett, s2s, short, w.get('source', '')))
    rows.sort(key=lambda x: -x[2])
    for i, (ef, et, amt, dt, eft, ett, s2s, flow, src) in enumerate(rows):
        r = i + 2
        ws.cell(r, 1, ef); ws.cell(r, 2, et)
        ws.cell(r, 3, amt).number_format = MONEY
        ws.cell(r, 4, dt or ''); ws.cell(r, 5, eft).alignment = CTR
        ws.cell(r, 6, ett).alignment = CTR; ws.cell(r, 7, s2s).alignment = CTR
        ws.cell(r, 8, flow); ws.cell(r, 9, src)
        fill = Z1 if i % 2 == 0 else Z2
        s2s_fill = PatternFill('solid', fgColor='FCE4EC') if s2s else fill
        for c in range(1, 10):
            ws.cell(r, c).fill = s2s_fill if c == 7 else fill; ws.cell(r, c).border = TB; ws.cell(r, c).font = BF
    widths(ws, [38, 38, 16, 12, 16, 16, 14, 24, 18])
    add_disclaimer(ws, 9)
    s2s_ct = sum(1 for x in rows if x[6] == 'YES')
    print(f"    {len(rows)} shell-involved, {s2s_ct} shell-to-shell")


# ══════════════════════════════════════════════════════════════
# TAB 10: SAR COMPARISON
# ══════════════════════════════════════════════════════════════
def tab_sar_comparison(wb):
    print("  Tab 10: SAR Comparison...")
    ws = wb.create_sheet("SAR Comparison")
    ws.sheet_properties.tabColor = "2C3E50"
    cols = ["Bank", "SAR Suspicious Activity (USD)", "Our Extraction (USD)",
            "Coverage %", "Source Authority", "Status"]
    make_header(ws, cols, DARK_HDR)
    r = 2
    for bank, data in SAR_BENCHMARKS.items():
        ws.cell(r, 1, bank).font = BOLD_DATA
        ws.cell(r, 2, data['amount']).number_format = MONEY
        ws.cell(r, 3, '—').alignment = CTR; ws.cell(r, 4, '—').alignment = CTR
        ws.cell(r, 5, data['source']).font = SF
        ws.cell(r, 6, 'Included in combined total').font = SF
        zebra(ws, r, 6); r += 1
    ws.cell(r, 1, "COMBINED BENCHMARK").font = Font(bold=True, name='Arial', size=11)
    ws.cell(r, 2, SAR_TOTAL).number_format = MONEY; ws.cell(r, 2).font = Font(bold=True, name='Arial', size=11)
    ws.cell(r, 3, GRAND_TOTAL).number_format = MONEY
    ws.cell(r, 3).font = Font(bold=True, name='Arial', size=11, color='1B6B3A')
    ws.cell(r, 4, GRAND_TOTAL / SAR_TOTAL).number_format = PCT; ws.cell(r, 4).font = GREEN_NUM
    ws.cell(r, 5, "All banks combined"); ws.cell(r, 6, "EXCEEDED ✓").font = GREEN_NUM
    for c in range(1, 7):
        ws.cell(r, c).fill = TOTAL_FILL; ws.cell(r, c).border = Border(top=Side(style='medium', color='1B2A4A'))
    r += 2
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1, "CALCULATION BREAKDOWN — How We Reached $1.964B").font = SUBTITLE; r += 1
    calc = [
        ("v2-20 Base: Amount-unique OCR extraction across 19 datasets", V2_20_BASE, "2,532 unique dollar amounts"),
        ("+ Phase 23: Same-amount different-date wire recovery", DATE_RECOVERED, "95 same-amount wires on new dates"),
        ("+ Phase 24: New amounts from master ledger tables", ABOVE_CAP - ABOVE_CAP + (GRAND_TOTAL - TIER_CONSERVATIVE - ABOVE_CAP), "32 amounts not seen in prior phases"),
        ("+ Phase 24: Above-cap court-verified wires", ABOVE_CAP, "8 wires verified with exhibit numbers"),
    ]
    for label, amt, note in calc:
        ws.cell(r, 1, label).font = BF; ws.cell(r, 2, amt).number_format = MONEY
        ws.cell(r, 3, note).font = SF; zebra(ws, r, 3); r += 1
    ws.cell(r, 1, "PUBLICATION TOTAL").font = Font(bold=True, name='Arial', size=12)
    ws.cell(r, 2, GRAND_TOTAL).number_format = MONEY
    ws.cell(r, 2).font = Font(bold=True, name='Arial', size=12, color='1B6B3A')
    ws.cell(r, 3, f"{GRAND_TOTAL/SAR_TOTAL*100:.1f}% of SAR benchmark").font = GREEN_NUM
    for c in range(1, 4):
        ws.cell(r, c).fill = TOTAL_FILL; ws.cell(r, c).border = Border(top=Side(style='medium', color='1B2A4A'))
    widths(ws, [50, 22, 22, 14, 32, 22])
    add_disclaimer(ws, 6)


# ══════════════════════════════════════════════════════════════
# TAB 11: METHODOLOGY
# ══════════════════════════════════════════════════════════════
def tab_methodology(wb):
    print("  Tab 11: Methodology & Limitations...")
    ws = wb.create_sheet("Methodology")
    ws.sheet_properties.tabColor = "7F8C8D"
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, "METHODOLOGY, QUALITY ASSURANCE & LIMITATIONS").font = TITLE_FONT
    ws.merge_cells('A2:E2')
    ws.cell(2, 1, "24-phase extraction pipeline · 9 bugs caught · 3-tier confidence framework"
            ).font = Font(name='Arial', size=10, color='7F8C8D', italic=True)

    r = 4
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(r, 1, "CONTAMINATION BUGS FOUND & CORRECTED").font = SUBTITLE; r += 1
    for c, val in enumerate(["Phase", "Bug Name", "Dollar Impact", "What Went Wrong", "How We Fixed It"], 1):
        cl = ws.cell(r, c, val); cl.font = HF; cl.fill = RED_HDR; cl.alignment = CTR
    r += 1
    for phase, name, impact, what, fix in BUGS_FOUND:
        ws.cell(r, 1, phase).alignment = CTR; ws.cell(r, 2, name).font = BOLD_DATA
        ws.cell(r, 3, impact).alignment = CTR; ws.cell(r, 4, what).font = SF; ws.cell(r, 5, fix).font = SF
        ws.row_dimensions[r].height = 48
        is_neg = impact.startswith('-'); is_pos = impact.startswith('+')
        fill = BUG_NEG if is_neg else (BUG_POS if is_pos else Z1)
        for c in range(1, 6):
            ws.cell(r, c).fill = fill; ws.cell(r, c).border = TB; ws.cell(r, c).alignment = LEFT_WRAP
        ws.cell(r, 1).alignment = CTR; ws.cell(r, 3).alignment = CTR
        ws.cell(r, 3).font = RED_NUM if is_neg else (GREEN_NUM if is_pos else BF)
        r += 1

    r += 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(r, 1, "LIMITATIONS & CAVEATS").font = SUBTITLE; r += 1
    for lim in [
        "This is automated text extraction, NOT a professional audit under GAAS/GAGAS.",
        "OCR quality varies. Some amounts may be misread (e.g., $1M vs $10M from scan artifacts).",
        "Entity names are normalized from OCR. Some entities may be misidentified or merged.",
        "Sealed/withheld documents (~$40-60M estimated) are not accessible.",
        "The SAR benchmark ($1.878B) counts attempted transactions; we find completed ones.",
        "Pre-retention period records may have been destroyed (~$15-30M estimated).",
        "Cross-table name duplication risk: same wire with different entity name formatting.",
        "Chain-hop filtering may occasionally exclude legitimate multi-step transactions.",
        "WEAK/VERY_WEAK tiers ($991M raw) were excluded — potential $5-15M additional.",
        "122 of 382 ledger entries have dates. Undated entries have higher duplication risk.",
    ]:
        ws.merge_cells(f'A{r}:E{r}'); ws.cell(r, 1, f"• {lim}").font = SF
        ws.cell(r, 1).alignment = LEFT_WRAP; ws.row_dimensions[r].height = 28; r += 1
    widths(ws, [14, 18, 14, 52, 52])


# ══════════════════════════════════════════════════════════════
# TAB 12: MULTI-BANK WIRE LEDGER (DB-driven)
# ══════════════════════════════════════════════════════════════
def tab_multi_bank_wires(wb, cur):
    print("  Tab 12: Multi-Bank Wire Ledger...")
    ws = wb.create_sheet("Multi-Bank Wires")

    # Get verified_wires
    cur.execute("""
        SELECT 'DB-SDNY' as source_tier, entity_from, entity_to, amount, date,
               bates, exhibit, 'wire_transfer' as tx_type, 'Deutsche Bank' as bank
        FROM verified_wires ORDER BY amount DESC
    """)
    verified = [dict(r) for r in cur.fetchall()]

    # Get bank_statement wire/book transfers
    cur.execute("""
        SELECT 'STATEMENT' as source_tier, '' as entity_from, '' as entity_to,
               tx_amount as amount, tx_date_raw as date,
               bates_primary as bates, '' as exhibit, tx_type, bank
        FROM bank_statement_transactions
        WHERE tx_type IN ('wire_transfer', 'book_transfer')
        ORDER BY tx_amount DESC
    """)
    statements = [dict(r) for r in cur.fetchall()]

    cols = ["Source Tier", "Bank", "Date", "Entity From", "Entity To",
            "Amount", "Type", "Bates / Exhibit"]
    make_header(ws, cols, TEAL_HDR)

    all_wires = verified + statements
    total = 0
    for r_idx, w in enumerate(sorted(all_wires, key=lambda x: -float(x['amount'] or 0)), 2):
        amt = float(w['amount'] or 0)
        total += amt
        tier = w['source_tier']
        ref = w.get('exhibit', '') or w.get('bates', '') or ''

        ws.cell(r_idx, 1, tier)
        ws.cell(r_idx, 2, sanitize(w.get('bank', '')))
        ws.cell(r_idx, 3, sanitize(w.get('date', '')))
        ws.cell(r_idx, 4, sanitize(w.get('entity_from', '')))
        ws.cell(r_idx, 5, sanitize(w.get('entity_to', '')))
        ws.cell(r_idx, 6, amt).number_format = MONEY
        ws.cell(r_idx, 7, sanitize(w.get('tx_type', '')))
        ws.cell(r_idx, 8, sanitize(ref))

        fill = PROVEN_FILL if tier == 'DB-SDNY' else Z1 if r_idx % 2 == 0 else Z2
        for c in range(1, 9):
            ws.cell(r_idx, c).fill = fill
            ws.cell(r_idx, c).font = BF
            ws.cell(r_idx, c).border = TB

    print(f"    {len(verified)} verified + {len(statements)} statement = {len(all_wires)} total wires")
    widths(ws, [12, 16, 14, 28, 28, 16, 14, 24])
    add_disclaimer(ws, 8)


# ══════════════════════════════════════════════════════════════
# TAB 13: BANK STATEMENT COVERAGE (DB-driven)
# ══════════════════════════════════════════════════════════════
def tab_bank_stmt_coverage(wb, cur):
    print("  Tab 13: Bank Statement Coverage...")
    ws = wb.create_sheet("Bank Stmt Coverage")

    # Get bank × year matrix
    cur.execute("""
        SELECT bank, year, COUNT(*) as cnt, SUM(tx_amount) as total
        FROM bank_statement_transactions
        WHERE bank IS NOT NULL AND year IS NOT NULL
        GROUP BY bank, year
        ORDER BY bank, year
    """)
    rows = cur.fetchall()

    banks = sorted(set(r['bank'] for r in rows))
    years = sorted(set(r['year'] for r in rows))

    # Header
    header = ["Bank"] + [str(y) for y in years] + ["TOTAL"]
    make_header(ws, header, PURPLE_HDR)

    for b_idx, bank in enumerate(banks, 2):
        ws.cell(b_idx, 1, bank).font = BOLD_DATA
        row_total = 0
        for y_idx, year in enumerate(years, 2):
            cnt = 0
            for r in rows:
                if r['bank'] == bank and r['year'] == year:
                    cnt = r['cnt']
                    break
            if cnt > 0:
                cl = ws.cell(b_idx, y_idx, cnt)
                # Heat map coloring
                if cnt >= 100:
                    cl.fill = PatternFill('solid', fgColor='1B6B3A')
                    cl.font = Font(name='Arial', size=9, color='FFFFFF', bold=True)
                elif cnt >= 50:
                    cl.fill = PatternFill('solid', fgColor='27AE60')
                    cl.font = Font(name='Arial', size=9, color='FFFFFF')
                elif cnt >= 10:
                    cl.fill = PatternFill('solid', fgColor='82E0AA')
                    cl.font = Font(name='Arial', size=9)
                elif cnt >= 1:
                    cl.fill = PatternFill('solid', fgColor='D5F5E3')
                    cl.font = Font(name='Arial', size=9)
                row_total += cnt
            else:
                cl = ws.cell(b_idx, y_idx, "")
                cl.fill = PatternFill('solid', fgColor='F2F2F2')
            cl.alignment = CTR
            cl.border = TB

        ws.cell(b_idx, len(years) + 2, row_total).font = BOLD_DATA
        ws.cell(b_idx, len(years) + 2).number_format = '#,##0'

    print(f"    {len(banks)} banks × {len(years)} years")
    widths(ws, [20] + [8] * len(years) + [10])


# ══════════════════════════════════════════════════════════════
# TAB 14: MULTI-BANK SUMMARY (DB-driven)
# ══════════════════════════════════════════════════════════════
def tab_multi_bank_summary(wb, cur, ledger):
    print("  Tab 14: Multi-Bank Summary...")
    ws = wb.create_sheet("Multi-Bank Summary")

    # ── Section 1: Pipeline comparison ──
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, "PIPELINE COMPARISON").font = TITLE_FONT

    headers1 = ["Source", "Records", "Amount", "Banks", "Date Range"]
    for c, h in enumerate(headers1, 1):
        cl = ws.cell(2, c, h); cl.font = HF; cl.fill = DARK_HDR; cl.alignment = CTR

    # master_wire_ledger
    mwl_total = sum(w['amount'] for w in ledger)
    ws.cell(3, 1, "Master Wire Ledger (Published)").font = BOLD_DATA
    ws.cell(3, 2, len(ledger)); ws.cell(3, 3, mwl_total).number_format = MONEY

    # bank_statement_transactions
    cur.execute("SELECT COUNT(*), SUM(tx_amount), COUNT(DISTINCT bank) FROM bank_statement_transactions")
    bst = cur.fetchone()
    ws.cell(4, 1, "Bank Statement Transactions").font = BF
    ws.cell(4, 2, bst[0]); ws.cell(4, 3, float(bst[1] or 0)).number_format = MONEY
    ws.cell(4, 4, bst[2])

    # verified_wires
    cur.execute("SELECT COUNT(*), SUM(amount) FROM verified_wires")
    vw = cur.fetchone()
    ws.cell(5, 1, "Verified Wires (Court Exhibits)").font = BF
    ws.cell(5, 2, vw[0]); ws.cell(5, 3, float(vw[1] or 0)).number_format = MONEY

    # fincen
    cur.execute("SELECT COUNT(*) FROM fincen_transactions")
    fc = cur.fetchone()
    ws.cell(6, 1, "FinCEN SAR Transactions").font = BF
    ws.cell(6, 2, fc[0])

    for r in range(3, 7):
        for c in range(1, 6):
            ws.cell(r, c).border = TB
            ws.cell(r, c).fill = Z1 if r % 2 == 0 else Z2

    # ── Section 2: Validation tiers ──
    r = 8
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(r, 1, "VALIDATION TIER BREAKDOWN").font = TITLE_FONT
    r += 1

    headers2 = ["Tier", "Records", "Amount", "% of Total", "Description"]
    for c, h in enumerate(headers2, 1):
        cl = ws.cell(r, c, h); cl.font = HF; cl.fill = GREEN_HDR; cl.alignment = CTR
    r += 1

    try:
        cur.execute("""
            SELECT validation_tier, COUNT(*), SUM(tx_amount)
            FROM bank_statement_transactions
            WHERE validation_tier IS NOT NULL
            GROUP BY validation_tier
            ORDER BY SUM(tx_amount) DESC
        """)
        tiers = cur.fetchall()
        bst_total = float(bst[1] or 1)

        tier_descs = {
            'FINCEN_VALIDATED': 'Amount + bank match in FinCEN SAR database',
            'FINCEN_AMOUNT_MATCH': 'Exact amount match in FinCEN (different bank)',
            'FINCEN_RANGE_MATCH': 'Amount within ±5% of a FinCEN record',
            'STATEMENT_ONLY': 'Bank statement only — no FinCEN corroboration',
        }

        for t in tiers:
            tier_name = t[0] or 'UNKNOWN'
            cnt = t[1]
            amt = float(t[2] or 0)
            ws.cell(r, 1, tier_name).font = BF
            ws.cell(r, 2, cnt)
            ws.cell(r, 3, amt).number_format = MONEY
            ws.cell(r, 4, amt / bst_total).number_format = PCT
            ws.cell(r, 5, tier_descs.get(tier_name, '')).font = SF
            for c in range(1, 6):
                ws.cell(r, c).border = TB
                ws.cell(r, c).fill = Z1 if r % 2 == 0 else Z2
            r += 1
    except Exception as e:
        ws.cell(r, 1, f"Validation tiers not available: {e}").font = SF
        r += 1

    # ── Section 3: Bank breakdown ──
    r += 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(r, 1, "BANK BREAKDOWN — ALL STATEMENT TRANSACTIONS").font = TITLE_FONT
    r += 1

    headers3 = ["Bank", "Records", "Amount", "Wire/Book", "Net-New"]
    for c, h in enumerate(headers3, 1):
        cl = ws.cell(r, c, h); cl.font = HF; cl.fill = TEAL_HDR; cl.alignment = CTR
    r += 1

    cur.execute("""
        SELECT bank, COUNT(*) as cnt, SUM(tx_amount) as total,
               SUM(CASE WHEN tx_type IN ('wire_transfer','book_transfer') THEN 1 ELSE 0 END) as wire_cnt,
               SUM(CASE WHEN dedup_status = 'NET_NEW' THEN 1 ELSE 0 END) as new_cnt
        FROM bank_statement_transactions
        GROUP BY bank
        ORDER BY SUM(tx_amount) DESC
    """)
    for row in cur.fetchall():
        ws.cell(r, 1, row['bank']).font = BF
        ws.cell(r, 2, row['cnt'])
        ws.cell(r, 3, float(row['total'] or 0)).number_format = MONEY
        ws.cell(r, 4, row['wire_cnt'])
        ws.cell(r, 5, row['new_cnt'])
        for c in range(1, 6):
            ws.cell(r, c).border = TB
            ws.cell(r, c).fill = Z1 if r % 2 == 0 else Z2
        r += 1

    # ── Section 4: 2003-2004 gap analysis ──
    r += 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(r, 1, "2003-2004 GAP ANALYSIS").font = TITLE_FONT
    r += 1

    cur.execute("""
        SELECT bank, COUNT(*) as cnt, SUM(tx_amount) as total
        FROM bank_statement_transactions
        WHERE year BETWEEN 2003 AND 2004
        GROUP BY bank
        ORDER BY SUM(tx_amount) DESC
    """)
    gap_rows = cur.fetchall()
    gap_total_cnt = sum(row['cnt'] for row in gap_rows)
    gap_total_amt = sum(float(row['total'] or 0) for row in gap_rows)

    ws.cell(r, 1, f"Total 2003-2004: {gap_total_cnt:,} transactions, ${gap_total_amt:,.0f}").font = BOLD_DATA
    ws.merge_cells(f'A{r}:E{r}')
    r += 1

    headers4 = ["Bank", "Transactions", "Amount", "", ""]
    for c, h in enumerate(headers4, 1):
        cl = ws.cell(r, c, h); cl.font = HF; cl.fill = RED_HDR; cl.alignment = CTR
    r += 1

    for row in gap_rows:
        ws.cell(r, 1, row['bank']).font = BF
        ws.cell(r, 2, row['cnt'])
        ws.cell(r, 3, float(row['total'] or 0)).number_format = MONEY
        for c in range(1, 4):
            ws.cell(r, c).border = TB
        r += 1

    print(f"    2003-04 gap: {gap_total_cnt} transactions, ${gap_total_amt:,.0f}")
    widths(ws, [28, 12, 18, 14, 40])
    add_disclaimer(ws, 5)


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
def main():
    print("=" * 70)
    print("  BUILDING FORENSIC WORKBOOK v7")
    print("  DATABASE-DRIVEN — Single Source of Truth")
    print("=" * 70)
    print()
    ledger = load_master_ledger()
    if not ledger:
        print("\n  Cannot build without master ledger. Exiting."); sys.exit(1)

    total = sum(w['amount'] for w in ledger)
    print(f"  Ledger: {len(ledger)} wires, ${total:,.0f}")
    print(f"  Published pipeline total: ${GRAND_TOTAL:,.0f} ({GRAND_TOTAL/SAR_TOTAL*100:.1f}% SAR)")
    print(f"  Entity-resolved ledger: ${total:,.0f} ({total/SAR_TOTAL*100:.1f}% SAR)")

    types = Counter(classify_entity(w.get('entity_from', '')) for w in ledger)
    types.update(classify_entity(w.get('entity_to', '')) for w in ledger)
    print(f"  Entity types: {dict(types)}")
    print()

    wb = Workbook()
    tab_executive_summary(wb)
    tab_extraction_phases(wb)
    tab_money_flow_patterns(wb, ledger)
    tab_shell_hierarchy(wb, ledger)
    tab_master_ledger(wb, ledger)
    tab_above_cap(wb, ledger)
    tab_date_recovery(wb, ledger)
    tab_entity_pl(wb, ledger)
    tab_shell_network(wb, ledger)
    tab_sar_comparison(wb)
    tab_methodology(wb)

    # ── NEW TABS 12-14 (database-driven) ──
    db_path = find_db()
    if db_path:
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            tab_multi_bank_wires(wb, cur)
            tab_bank_stmt_coverage(wb, cur)
            tab_multi_bank_summary(wb, cur, ledger)
            conn.close()
        except Exception as e:
            print(f"  ⚠ New tabs skipped: {e}")
    else:
        print("  ⚠ No database found — tabs 12-14 skipped")

    local_path = REPORT_DIR / OUTFILE
    wb.save(local_path)
    print(f"\n  Saved: {local_path}")
    print(f"  Size: {local_path.stat().st_size // 1024} KB")
    if HDD_DIR.exists():
        import shutil
        shutil.copy2(local_path, HDD_DIR / OUTFILE)
        print(f"  Copied to HDD: {HDD_DIR / OUTFILE}")
        shutil.copy2(local_path, Path.cwd() / OUTFILE)
        print(f"  Copied to CWD: {Path.cwd() / OUTFILE}")

    tab_count = len(wb.sheetnames)
    print(f"\n{'=' * 70}")
    print(f"  FORENSIC WORKBOOK v7 COMPLETE — {tab_count} TABS")
    for i, name in enumerate(wb.sheetnames, 1):
        marker = " ★ NEW" if i > 11 else ""
        print(f"    Tab {i:2d}: {name}{marker}")
    print(f"  Published pipeline: ${GRAND_TOTAL:,.0f} ({GRAND_TOTAL/SAR_TOTAL*100:.1f}% SAR)")
    print(f"  Entity-resolved ledger: {len(ledger)} wires, ${sum(w['amount'] for w in ledger):,.0f}")
    print(f"{'=' * 70}")

if __name__ == "__main__":
    main()
